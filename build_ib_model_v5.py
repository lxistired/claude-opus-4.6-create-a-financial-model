"""
Alphabet Inc. (GOOG) - IB Financial Model v5
=============================================
11 Tabs: Cover | Key_Summary | Assumptions | Segment_Revenue | Segment_PL
         | Consolidated_PL | BS | Cash_Flow | DCF | Sensitivity | Ratio_Analysis

v5 upgrade (vs v4):
  - NEW: Key_Summary executive dashboard (KPIs + investment highlights)
  - NEW: Balance Sheet (full 3-statement linkage)
  - NEW: Ratio_Analysis (profitability / efficiency / leverage / growth)
  - ENHANCED: Notes column (col J) on every data tab explaining prediction logic
  - ENHANCED: Data legend + sources at bottom of every tab
  - ENHANCED: Cash_Flow with Working Capital detail
  - FIX: 2023 Cash & Securities corrected to ~$110.9B (was duplicated from 2024)
  - Source: Alphabet 2025 10-K (filed Feb 4 2026), pp.48-52 for financial statements
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter as CL
import os

# ═══════════════════════════════════════════════════════
#  STYLES
# ═══════════════════════════════════════════════════════
BLUE = PatternFill("solid", fgColor="DDEBF7")
HDR  = PatternFill("solid", fgColor="1F4E78")
SUB  = PatternFill("solid", fgColor="D9E1F2")
KEY  = PatternFill("solid", fgColor="FFF2CC")
GREY = PatternFill("solid", fgColor="F2F2F2")
ERR  = PatternFill("solid", fgColor="FFC7CE")
FW   = Font(color="FFFFFF", bold=True, size=10)
FB   = Font(bold=True, size=10)
FI   = Font(color="1F4E78", bold=True, size=10)          # assumption input
FH   = Font(color="305496", size=10)                      # historical hardcoded
FL   = Font(color="548235", size=10)                       # link from other sheet
FN   = Font(size=10)                                       # normal / formula
FT   = Font(size=14, bold=True)
FS   = Font(size=11, bold=True, color="1F4E78")
FSM  = Font(size=9, italic=True, color="808080")
FNOTE = Font(size=9, color="808080")
TH   = Side(style="thin", color="D9D9D9")
BD   = Border(left=TH, right=TH, top=TH, bottom=TH)
CT   = Alignment(horizontal="center", vertical="center")
LT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
NI   = "#,##0"
NP   = "0.0%"
NP2  = "0.00%"
NU   = "$#,##0.00"
YR   = [2023, 2024, 2025, "2026E", "2027E", "2028E", "2029E", "2030E"]
NCOL = 11  # Notes column (K), avoids conflict with 2030E data in col J

# ═══════════════════════════════════════════════════════
#  HISTORICAL DATA (from 10-K pp.48-52, Note 15)
# ═══════════════════════════════════════════════════════
# Revenue by segment ($mm)
SEG_REV = {
    "Google Search & other":      [175033, 198084, 224532],
    "YouTube ads":                [31510, 36147, 40367],
    "Google Network":             [31312, 30359, 29792],
    "Subs, Platforms & Devices":  [34688, 40340, 48030],
    "Google Cloud":               [33088, 43229, 58705],
    "Other Bets":                 [1527, 1648, 1537],
}
HEDGE = [236, 211, -127]

# Consolidated P&L ($mm) - from p.49
PL = {
    "Revenue":    [307394, 350018, 402836],
    "COGS":       [133332, 146306, 162535],
    "TAC":        [49316, 54900, 59926],
    "OtherCOGS":  [84016, 91406, 102609],
    "R&D":        [45427, 49326, 61087],
    "S&M":        [27917, 27808, 28693],
    "G&A":        [16425, 14188, 21482],
    "EBIT":       [84293, 112390, 129039],
    "OIE":        [1424, 7425, 29787],
    "Tax":        [11922, 19697, 26656],
    "NI":         [73795, 100118, 132170],
    "D&A":        [11946, 15311, 21136],
    "SBC":        [22460, 22785, 24953],
}
EPS_D = [5.80, 8.04, 10.81]
SHARES = [12722, 12447, 12230]

# Segment P&L ($mm) - from Note 15
SEG_PL = {
    "svc_emp": [46224, 44560, 45124],
    "svc_oth": [130461, 139107, 158193],
    "svc_oi":  [95858, 121263, 139404],
    "cld_emp": [19054, 20519, 22078],
    "cld_oth": [12318, 16598, 22717],
    "cld_oi":  [1716, 6112, 13910],
    "ob_oi":   [-4095, -4444, -7515],
    "alpha":   [-9186, -10541, -16760],
}

# Balance Sheet ($mm) - from p.48 (2024 & 2025 exact; 2023 estimated from prior filing)
BS_DATA = {
    "Cash":              [24048, 23466, 30708],
    "MktSec":            [86868, 72191, 96135],
    "CashTotal":         [110916, 95657, 126843],
    "AR":                [40300, 52340, 62886],
    "OtherCA":           [12200, 15714, 16309],
    "TotalCA":           [163416, 163711, 206038],
    "NonMktSec":         [31008, 37982, 68687],
    "DeferredTax":       [12270, 17180, 9113],
    "PPE":               [134345, 171036, 246597],
    "OpLeaseAsset":      [14253, 13588, 15221],
    "Goodwill":          [29198, 31885, 33380],
    "OtherNCA":          [16902, 14874, 16245],
    "TotalAssets":       [401392, 450256, 595281],
    "AP":                [7100, 7987, 12200],
    "AccComp":           [14200, 15069, 17546],
    "AccExp":            [43000, 51228, 55557],
    "RevShare":          [8743, 9802, 10864],
    "DefRev":            [3993, 5036, 6578],
    "TotalCL":           [77036, 89122, 102745],
    "LTDebt":            [12887, 10883, 46547],
    "TaxNC":             [8049, 8782, 9531],
    "OpLeaseLiab":       [10989, 11691, 12744],
    "OtherLTL":          [9052, 4694, 8449],
    "TotalLiab":         [118013, 125172, 180016],
    "Equity":            [283379, 325084, 415265],
}

# Cash Flow ($mm) - from pp.51-52
CF = {
    "NI":         [73795, 100118, 132170],
    "DA":         [11946, 15311, 21136],
    "SBC":        [22460, 22785, 24953],
    "DeferredTax":[-7763, -5257, 8348],
    "SecGL":      [823, -2671, -24620],
    "Other":      [4330, 3419, 2108],
    "ChgAR":      [-7833, -5891, -8779],
    "ChgTaxNet":  [523, -2418, -3226],
    "ChgOtherA":  [-2143, -1397, -4542],
    "ChgAP":      [664, 359, 907],
    "ChgAccExp":  [3937, -1161, 12939],
    "ChgRevShare":[482, 1059, 899],
    "ChgDefRev":  [525, 1043, 2420],
    "CFO":        [101746, 125299, 164713],
    "CapEx":      [-32251, -52535, -91447],
    "BuyMktSec":  [-77858, -86679, -103773],
    "SellMktSec": [86672, 103428, 83240],
    "OtherInv":   [-3626, -9750, -8311],
    "CFI":        [-27063, -45536, -120291],
    "StockNet":   [-9837, -12190, -14167],
    "Buyback":    [-61504, -62222, -45709],
    "Dividend":   [0, -7363, -10049],
    "DebtIssue":  [10790, 13589, 64564],
    "DebtRepay":  [-11550, -12701, -32427],
    "OtherFin":   [8, 1154, 400],
    "CFF":        [-72093, -79733, -37388],
    "FX":         [-421, -612, 208],
}

# ═══════════════════════════════════════════════════════
#  HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════
def apply_bd(ws, r, c):
    ws.cell(row=r, column=c).border = BD

def sb(ws, r1, c1, r2, c2, fill=None, font=None, fmt=None):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            cell = ws.cell(row=r, column=c); cell.border = BD
            if fill: cell.fill = fill
            if font: cell.font = font
            if fmt: cell.number_format = fmt

def cw(ws, d):
    for k, v in d.items(): ws.column_dimensions[k].width = v

def yh(ws, row, labels=None, ncols=9):
    labels = labels or YR
    for i, yr in enumerate(labels):
        c = ws.cell(row=row, column=2+i, value=yr)
        c.font = FW; c.fill = HDR; c.alignment = CT; c.border = BD
    ws.cell(row=row, column=1).fill = HDR; ws.cell(row=row, column=1).border = BD
    n = ws.cell(row=row, column=NCOL, value="Notes / 预测逻辑")
    n.font = FW; n.fill = HDR; n.alignment = CT; n.border = BD

def sec(ws, r, label, ncols=11):
    ws.cell(row=r, column=1, value=label).font = FS
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    for c in range(1, ncols+1):
        ws.cell(row=r, column=c).fill = SUB; ws.cell(row=r, column=c).border = BD

def note(ws, r, text):
    # Prevent text starting with '=' being treated as formula by openpyxl
    if text and text.startswith('='):
        text = text[1:]  # strip leading '='
    c = ws.cell(row=r, column=NCOL, value=text)
    c.font = FNOTE; c.alignment = LT; c.border = BD

def pct_row(ws, r, num_r, den_r, ncols=9):
    ws.cell(row=r, column=1).border = BD
    for c in range(2, ncols+1):
        ws.cell(row=r, column=c, value=f"={CL(c)}{num_r}/{CL(c)}{den_r}").number_format = NP
        ws.cell(row=r, column=c).border = BD; ws.cell(row=r, column=c).font = FSM

def hval(ws, r, col, val, fmt=NI):
    """Write historical value with blue font"""
    c = ws.cell(row=r, column=col, value=val)
    c.number_format = fmt; c.font = FH; c.border = BD

def fval(ws, r, col, formula, fmt=NI):
    """Write formula value"""
    c = ws.cell(row=r, column=col, value=formula)
    c.number_format = fmt; c.border = BD

def lval(ws, r, col, formula, fmt=NI):
    """Write link from other sheet with green font"""
    c = ws.cell(row=r, column=col, value=formula)
    c.number_format = fmt; c.font = FL; c.border = BD

def arow(ws, r, label, h23, h24, h25, base, bull, bear, fmt=NP, nt=""):
    """Assumption row: historical + Base/Bull/Bear + CHOOSE"""
    ws.cell(row=r, column=1, value=label).font = FB; ws.cell(row=r, column=1).border = BD
    for ci, v in [(2,h23),(3,h24),(4,h25)]:
        c = ws.cell(row=r, column=ci, value=v)
        c.fill = GREY; c.font = FH; c.number_format = fmt; c.border = BD
    for ci, v in [(6,base),(7,bull),(8,bear)]:
        c = ws.cell(row=r, column=ci, value=v)
        c.fill = BLUE; c.font = FI; c.number_format = fmt; c.border = BD
    sel = ws.cell(row=r, column=9,
                  value=f'=CHOOSE(MATCH($B$2,{{"Base","Bull","Bear"}},0),F{r},G{r},H{r})')
    sel.number_format = fmt; sel.border = BD; sel.font = FB
    if nt: note(ws, r, nt)

def add_legend(ws, row):
    """Add data legend and sources at bottom of sheet"""
    ws.cell(row=row, column=1, value="数据图例 Data Legend:").font = Font(bold=True, size=9, color="808080")
    ws.cell(row=row+1, column=1, value="  蓝色字体 Blue Font = 财报原始数据 Hardcoded from financials").font = FNOTE
    ws.cell(row=row+2, column=1, value="  绿色字体 Green Font = 链接自其他工作表 Links from other sheets").font = FNOTE
    ws.cell(row=row+3, column=1, value="  蓝底深蓝字体 Blue Fill = 可编辑假设 Editable assumption inputs").font = FNOTE
    ws.cell(row=row+4, column=1, value="  黄底粗体 Yellow Fill = 关键输出 Key output rows").font = FNOTE
    ws.cell(row=row+5, column=1, value="  黑色字体 Black Font = 公式计算 Calculated from formulas").font = FNOTE
    ws.cell(row=row+7, column=1,
            value="数据来源 Data Sources:").font = Font(bold=True, size=9, color="808080")
    ws.cell(row=row+8, column=1,
            value="  • Alphabet 2025 10-K Annual Report (filed Feb 4, 2026) - Consolidated Financial Statements pp.48-52").font = FNOTE
    ws.cell(row=row+9, column=1,
            value="  • Alphabet 2025 10-K Note 15: Segment Information (Google Services, Cloud, Other Bets)").font = FNOTE
    ws.cell(row=row+10, column=1,
            value="  • Q4 2025 Earnings Call (Feb 4, 2026): CapEx 2026 guidance $75B (single quarter) → annualized ~$180B").font = FNOTE
    ws.cell(row=row+11, column=1,
            value="  • 10Y UST yield ~4.2% (Feb 2026), Fed 2% long-term inflation target").font = FNOTE


def build():
    wb = Workbook()
    R = {}  # cross-tab row references

    # ═══════════════════════════════════════════════════
    #  TAB 1: COVER
    # ═══════════════════════════════════════════════════
    ws = wb.active; ws.title = "Cover"
    ws["A1"] = "Alphabet Inc. (GOOG/GOOGL)"; ws["A1"].font = Font(size=20, bold=True)
    ws["A2"] = "Segment-Driven DCF Model  |  v5  |  2026-02-10"; ws["A2"].font = Font(size=12, italic=True, color="1F4E78")
    info = [
        (4, "数据来源 Data Sources", "Alphabet 2025 10-K + Q4'25 Earnings Call + FRED"),
        (5, "CapEx 2026 指引", "$75B/quarter → annualized ~$180B (earnings call 2026-02-04)"),
        (6, "模型特色 Key Feature", "11-Tab model: 三表联动 (PL + BS + CF) + Segment P&L 交叉验证"),
        (7, "  ", "Dual-layer P&L: Segment_PL (by division) + Consolidated_PL (by cost type)"),
        (9, "使用步骤 HOW TO USE", ""),
        (10, "Step 1", "Go to Assumptions → edit BLUE cells only (蓝底单元格)"),
        (11, "Step 2", "Switch scenario: Assumptions!B2 → Base / Bull / Bear"),
        (12, "Step 3", "All tabs auto-update: Revenue → Segment PL → Consolidated PL → BS → CF → DCF"),
        (13, "Step 4", "Check Segment_PL cross-check row: difference should = 0"),
        (14, "Step 5", "Check BS balance check row: Total Assets = Total L&E"),
        (15, "Step 6", "Every assumption shows HISTORICAL values (2023-2025) for reference"),
        (16, "Step 7", "Notes column (col J) explains prediction logic for every row"),
    ]
    for r, a, b in info:
        ws.cell(row=r, column=1, value=a).font = FB
        if b: ws.cell(row=r, column=2, value=b)
    cw(ws, {"A": 28, "B": 90})

    # ═══════════════════════════════════════════════════
    #  TAB 2: KEY SUMMARY (NEW)
    # ═══════════════════════════════════════════════════
    ws = wb.create_sheet("Key_Summary")
    ws["A1"] = "Alphabet Inc. (GOOG) - 财务摘要 Executive Summary"; ws["A1"].font = FT
    ws["A2"] = "单位：百万美元（USD Million）"; ws["A2"].font = FSM
    cw(ws, {"A": 38, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 38})
    # We'll fill this after building other tabs so we can link
    # Store starting row for later fill
    R["ks_start"] = 4

    # ═══════════════════════════════════════════════════
    #  TAB 3: ASSUMPTIONS
    # ═══════════════════════════════════════════════════
    ws = wb.create_sheet("Assumptions")
    cw(ws, {"A": 44, "B": 14, "C": 14, "D": 14, "E": 3, "F": 14, "G": 14, "H": 14, "I": 14, "J": 14, "K": 40})

    ws["A1"] = "Assumptions & Drivers 假设与驱动"; ws["A1"].font = FT
    ws["A2"] = "Scenario 情景"; ws["B2"] = "Base"; ws["B2"].fill = BLUE; ws["B2"].font = FI
    dv = DataValidation(type="list", formula1='"Base,Bull,Bear"', allow_blank=False)
    ws.add_data_validation(dv); dv.add(ws["B2"])

    for i, h in enumerate(["Parameter 参数", "2023A", "2024A", "2025A", "", "Base", "Bull", "Bear", "Selected", "", "Notes / 预测逻辑"]):
        c = ws.cell(row=4, column=1+i, value=h)
        c.font = FW; c.fill = HDR; c.alignment = CT; c.border = BD

    # ── WACC ──
    sec(ws, 6, "WACC 加权平均资本成本参数")
    arow(ws, 7,  "Risk-Free Rate 无风险利率 (10Y UST)", None, None, None, 0.042, 0.042, 0.042, NP2, "10Y美债收益率 ~4.2% (Feb 2026)")
    arow(ws, 8,  "Equity Beta 权益Beta", None, None, None, 1.05, 1.00, 1.10, "0.00", "参考Bloomberg/Reuters 5Y weekly beta")
    arow(ws, 9,  "Equity Risk Premium ERP 股权风险溢价", None, None, None, 0.050, 0.052, 0.048, NP2, "Damodaran 2026 ERP estimate")
    arow(ws, 10, "Pre-tax Cost of Debt Kd 税前债务成本", None, None, None, 0.049, 0.048, 0.052, NP2, "参考Alphabet最新债券到期收益率")
    arow(ws, 11, "Effective Tax Rate 有效税率", 0.139, 0.164, 0.168, 0.170, 0.165, 0.180, NP, "2023-2025历史税率参考")
    arow(ws, 12, "Debt / Total Capital 债务占比", None, None, None, 0.030, 0.025, 0.035, NP2, "Alphabet资本结构以权益为主")
    arow(ws, 13, "Terminal Growth Rate g 永续增长率", None, None, None, 0.030, 0.033, 0.027, NP2, "约=长期GDP增长+通胀")

    sec(ws, 15, "Derived WACC 推导的WACC")
    for r, lbl, f in [(16, "Ke = Rf + Beta × ERP 股权成本", "=I7+I8*I9"),
                       (17, "After-tax Kd 税后债务成本", "=I10*(1-I11)"),
                       (18, "WACC 加权平均资本成本", "=(1-I12)*I16+I12*I17")]:
        ws.cell(row=r, column=1, value=lbl).font = FB; ws.cell(row=r, column=1).border = BD
        c = ws.cell(row=r, column=9, value=f); c.number_format = NP2; c.border = BD
        if r == 18: c.fill = KEY; c.font = FB

    # ── Valuation Bridge ──
    sec(ws, 20, "Valuation Bridge 估值桥接")
    arow(ws, 21, "Cash & Marketable Securities 现金及有价证券 (mm)", 110916, 95657, 126843, 126843, 126843, 126843, NI, "年末现金+短期证券合计")
    arow(ws, 22, "Total Debt 总债务 (mm)", 12887, 10883, 46547, 46547, 46547, 46547, NI, "短期+长期债务")
    arow(ws, 23, "Diluted Shares 稀释股数 (mm)", 12722, 12447, 12230, 12230, 12230, 12230, NI, "Source: 10-K p.49")

    # ── Revenue Growth ──
    sec(ws, 25, "Revenue Growth by Segment 分部收入增长率 (Base块)")
    seg_names = list(SEG_REV.keys())
    for i, yr in enumerate(["2023A", "2024A", "2025A", "", "2026E", "2027E", "2028E", "2029E", "2030E"]):
        c = ws.cell(row=26, column=2+i, value=yr if yr else "")
        if yr: c.font = FW; c.fill = HDR; c.alignment = CT; c.border = BD
    ws.cell(row=26, column=1, value="Segment").font = FW; ws.cell(row=26, column=1).fill = HDR; ws.cell(row=26, column=1).border = BD

    base_g = [[.12,.10,.09,.08,.07],[.14,.12,.10,.09,.08],[-.02,-.02,-.01,-.01,0],
              [.18,.15,.13,.11,.09],[.30,.25,.22,.18,.15],[.10,.15,.15,.15,.10]]
    bull_g = [[.15,.13,.11,.10,.09],[.18,.15,.13,.11,.10],[-.01,0,0,.01,.01],
              [.22,.18,.15,.13,.11],[.38,.30,.25,.20,.17],[.15,.20,.20,.20,.15]]
    bear_g = [[.08,.07,.06,.05,.04],[.08,.07,.06,.05,.04],[-.05,-.04,-.03,-.03,-.02],
              [.12,.10,.09,.08,.07],[.22,.18,.15,.12,.10],[0,.05,.05,.05,.05]]

    seg_base_rows = {}
    seg_notes = ["搜索广告核心引擎", "AI-driven Shorts增长", "持续缩减", "YouTube Premium/硬件增长",
                 "AI Cloud高增长引擎", "Waymo等前沿业务"]
    for idx, name in enumerate(seg_names):
        r = 27 + idx; seg_base_rows[name] = r
        ws.cell(row=r, column=1, value=name).font = FB; ws.cell(row=r, column=1).border = BD
        hist = SEG_REV[name]
        for j in [1,2]:
            g = hist[j]/hist[j-1]-1
            hval(ws, r, 2+j, g, NP)
        for j, v in enumerate(base_g[idx]):
            c = ws.cell(row=r, column=6+j, value=v); c.fill = BLUE; c.font = FI; c.number_format = NP; c.border = BD
        note(ws, r, seg_notes[idx])

    sec(ws, 34, "Revenue Growth - Bull Case 乐观情景")
    seg_bull_rows = {}
    for idx, name in enumerate(seg_names):
        r = 35 + idx; seg_bull_rows[name] = r
        ws.cell(row=r, column=1, value=name).font = FB; ws.cell(row=r, column=1).border = BD
        for j, v in enumerate(bull_g[idx]):
            c = ws.cell(row=r, column=6+j, value=v); c.fill = BLUE; c.font = FI; c.number_format = NP; c.border = BD

    sec(ws, 42, "Revenue Growth - Bear Case 悲观情景")
    seg_bear_rows = {}
    for idx, name in enumerate(seg_names):
        r = 43 + idx; seg_bear_rows[name] = r
        ws.cell(row=r, column=1, value=name).font = FB; ws.cell(row=r, column=1).border = BD
        for j, v in enumerate(bear_g[idx]):
            c = ws.cell(row=r, column=6+j, value=v); c.fill = BLUE; c.font = FI; c.number_format = NP; c.border = BD

    # ── Segment Cost Structure ──
    sec(ws, 50, "Segment Cost Structure 分部成本结构 (with Historical)")
    for i, h in enumerate(["Cost Item 成本项", "2023A", "2024A", "2025A", "", "Base", "Bull", "Bear", "Selected", "", "Notes"]):
        c = ws.cell(row=51, column=1+i, value=h); c.font = FW; c.fill = HDR; c.alignment = CT; c.border = BD
    arow(ws, 52, "Services: Emp Comp % 员工薪酬占比", .170, .146, .132, .128, .125, .135, NP, "薪酬含SBC+福利，Note15")
    arow(ws, 53, "Services: Other Costs % 其他费用占比", .479, .456, .461, .455, .448, .468, NP, "含TAC/内容采购/基础设施")
    arow(ws, 54, "Cloud: Emp Comp % 员工薪酬占比", .576, .475, .376, .340, .330, .360, NP, "Cloud人效持续改善")
    arow(ws, 55, "Cloud: Other Costs % 其他费用占比", .372, .384, .387, .380, .370, .395, NP, "基础设施和第三方服务")
    arow(ws, 56, "Other Bets: Op Margin 经营利润率", -2.682, -2.696, -4.889, -3.500, -3.000, -4.500, NP, "前沿业务持续亏损")
    arow(ws, 57, "Alphabet-level Costs 公司级费用 (mm)", 9186, 10541, 16760, 18000, 17000, 20000, NI, "未分配AI研发+企业费用")

    # ── Consolidated Cost ──
    sec(ws, 59, "Consolidated Cost Structure 合并成本率")
    for i, h in enumerate(["Cost Item", "2023A", "2024A", "2025A", "", "Base", "Bull", "Bear", "Selected", "", "Notes"]):
        c = ws.cell(row=60, column=1+i, value=h); c.font = FW; c.fill = HDR; c.alignment = CT; c.border = BD
    arow(ws, 61, "TAC Rate 流量获取成本率 (% Ad Rev)", .207, .207, .203, .200, .198, .208, NP, "支付给合作伙伴的分成")
    arow(ws, 62, "Other COGS 其他营业成本 (% Total Rev)", .273, .261, .255, .250, .245, .260, NP, "基础设施折旧/内容/数据中心")
    arow(ws, 63, "R&D 研发费用 (% Total Rev)", .148, .141, .152, .145, .140, .155, NP, "AI投入持续，但规模效应显现")
    arow(ws, 64, "S&M 销售与市场 (% Total Rev)", .091, .079, .071, .068, .065, .074, NP, "效率提升，销售费率下降")
    arow(ws, 65, "G&A 管理费用 (% Total Rev)", .053, .041, .053, .040, .038, .045, NP, "2025含一次性诉讼，正常化后下降")
    arow(ws, 66, "D&A 折旧摊销 (% Rev)", .039, .044, .052, .055, .053, .058, NP, "CapEx高增长推动D&A上升")
    arow(ws, 67, "SBC 股权激励 (% Rev)", .073, .065, .062, .058, .055, .063, NP, "占比稳步下降")
    arow(ws, 68, "NWC 净营运资本变动 (% incr Rev)", None, None, None, .005, .004, .007, NP, "科技公司NWC需求低")

    # ── CapEx ──
    sec(ws, 70, "CapEx 资本支出 (USD mm) - 2026 per mgmt guidance ~$180B")
    for i, h in enumerate(["Year", "2023A", "2024A", "2025A", "", "Base", "Bull", "Bear", "Selected", "", "Notes"]):
        c = ws.cell(row=71, column=1+i, value=h); c.font = FW; c.fill = HDR; c.alignment = CT; c.border = BD
    capex_data = [
        (72, "2026E CapEx", 32251, 52535, 91447, 180000, 185000, 175000, "Q4'25 Earnings: $75B/Q → ~$180B/yr"),
        (73, "2027E CapEx", None, None, None, 160000, 170000, 155000, "AI投资高峰后逐步回落"),
        (74, "2028E CapEx", None, None, None, 140000, 150000, 135000, "资本密集度开始降低"),
        (75, "2029E CapEx", None, None, None, 120000, 130000, 115000, "CapEx周期转为维护型"),
        (76, "2030E CapEx", None, None, None, 105000, 115000, 100000, "稳态CapEx水平"),
    ]
    capex_rows = []
    for r_num, label, h23, h24, h25, base, bull, bear, nt in capex_data:
        arow(ws, r_num, label, h23, h24, h25, base, bull, bear, NI, nt)
        capex_rows.append(r_num)

    # ── Capital Return ──
    sec(ws, 78, "Capital Return 资本回报")
    arow(ws, 79, "Annual Buyback 年度回购 (mm)", 62184, 62047, 45398, 50000, 55000, 40000, NI, "2025因发债减少回购，2026恢复")
    arow(ws, 80, "Dividend per Share 每股股息 (annual)", 0, 0.60, 0.83, 0.88, 0.92, 0.84, NU, "2024年首次派息$0.60/share")

    # ── BS Assumptions (NEW) ──
    sec(ws, 82, "Balance Sheet Assumptions 资产负债表假设")
    for i, h in enumerate(["Item", "2023A", "2024A", "2025A", "", "Base", "Bull", "Bear", "Selected", "", "Notes"]):
        c = ws.cell(row=83, column=1+i, value=h); c.font = FW; c.fill = HDR; c.alignment = CT; c.border = BD
    arow(ws, 84, "AR Days 应收账款周转天数", 47.8, 54.6, 57.0, 55, 53, 58, "0.0", "AR/Revenue×365")
    arow(ws, 85, "AP Days 应付账款周转天数 (of COGS)", 19.4, 19.9, 27.4, 26, 27, 24, "0.0", "AP/COGS×365")
    arow(ws, 86, "AccComp 应计薪酬 (% Rev)", .046, .043, .044, .043, .042, .045, NP, "员工薪酬应计占收入比")
    arow(ws, 87, "AccExp+Other 应计费用 (% Rev)", .140, .146, .138, .135, .130, .142, NP, "运营应计费用占收入比")
    arow(ws, 88, "RevShare 应计收入分成 (% Ad Rev)", .037, .037, .037, .036, .035, .038, NP, "广告收入分成应计")
    arow(ws, 89, "DefRev 递延收入增速 %", None, .261, .306, .15, .18, .12, NP, "Cloud预付合同增长")
    arow(ws, 90, "OpLease 经营租赁资产增速 %", None, -.047, .120, .05, .06, .03, NP, "办公/数据中心租赁")
    arow(ws, 91, "NonMktSec 非上市证券增速 %", None, .225, .808, .15, .20, .10, NP, "2025因估值调增大幅增长")

    ws.freeze_panes = "B5"
    add_legend(ws, 95)

    # ═══════════════════════════════════════════════════
    #  TAB 4: SEGMENT REVENUE
    # ═══════════════════════════════════════════════════
    ws = wb.create_sheet("Segment_Revenue")
    ws["A1"] = "Revenue Build 分部收入构建 (USD mm)"; ws["A1"].font = FT
    ws["A2"] = "单位：百万美元 | Cloud Q4'25 = $17.7B (48% YoY)"; ws["A2"].font = FSM
    cw(ws, {CL(c): (40 if c == 1 else (38 if c == NCOL else 15)) for c in range(1, NCOL+1)})
    yh(ws, 4)

    def gf(seg_idx, yr_j):
        """CHOOSE formula for growth from 3 scenario blocks"""
        yc = CL(6+yr_j)
        br = seg_base_rows[seg_names[seg_idx]]
        bur = seg_bull_rows[seg_names[seg_idx]]
        ber = seg_bear_rows[seg_names[seg_idx]]
        return f'CHOOSE(MATCH(Assumptions!$B$2,{{"Base","Bull","Bear"}},0),Assumptions!{yc}${br},Assumptions!{yc}${bur},Assumptions!{yc}${ber})'

    r = 6
    # Google Advertising
    sec(ws, r, "Google Advertising 谷歌广告"); r += 1
    rev_rows = {}
    seg_notes_rev = {
        "Google Search & other": "搜索广告核心收入，AI Overview推动",
        "YouTube ads": "Shorts变现+Connected TV增长",
        "Google Network": "持续萎缩，向自有平台转移",
        "Subs, Platforms & Devices": "YouTube Premium/Music + Pixel设备",
        "Google Cloud": "GCP+Workspace，AI Cloud核心增长引擎",
        "Other Bets": "Waymo/Verily等前沿业务",
    }
    for seg_idx in range(3):
        name = seg_names[seg_idx]
        ws.cell(row=r, column=1, value=name).font = FB; ws.cell(row=r, column=1).border = BD
        rev_rows[name] = r
        for j, v in enumerate(SEG_REV[name]):
            hval(ws, r, 2+j, v)
        for yr_j in range(5):
            pc = 5+yr_j; prev = "D" if yr_j == 0 else CL(pc-1)
            fval(ws, r, pc, f"={prev}{r}*(1+{gf(seg_idx, yr_j)})")
        note(ws, r, seg_notes_rev[name])
        r += 1
        # yoy
        ws.cell(row=r, column=1, value="  yoy %").font = FSM; ws.cell(row=r, column=1).border = BD
        for j in [1,2]:
            fval(ws, r, 2+j, f"={CL(2+j)}{r-1}/{CL(1+j)}{r-1}-1", NP)
            ws.cell(row=r, column=2+j).font = FSM
        for yr_j in range(5):
            pc = 5+yr_j
            fval(ws, r, pc, f"={CL(pc)}{r-1}/{CL(pc-1)}{r-1}-1", NP)
            ws.cell(row=r, column=pc).font = FSM
        r += 1

    ad_total_r = r
    ws.cell(row=r, column=1, value="Total Google Advertising 广告合计").font = FB
    sb(ws, r, 1, r, 9, fill=SUB, font=FB)
    for c in range(2, 10):
        parts = "+".join(f"{CL(c)}{rev_rows[n]}" for n in seg_names[:3])
        fval(ws, r, c, f"={parts}")
    note(ws, r, "=Search+YouTube+Network")
    r += 2

    # Subs
    sec(ws, r, "Subscriptions, Platforms & Devices 订阅/平台/设备"); r += 1
    name = seg_names[3]
    ws.cell(row=r, column=1, value=name).font = FB; ws.cell(row=r, column=1).border = BD
    rev_rows[name] = r
    for j, v in enumerate(SEG_REV[name]):
        hval(ws, r, 2+j, v)
    for yr_j in range(5):
        pc = 5+yr_j; prev = "D" if yr_j == 0 else CL(pc-1)
        fval(ws, r, pc, f"={prev}{r}*(1+{gf(3, yr_j)})")
    note(ws, r, seg_notes_rev[name])
    r += 1
    ws.cell(row=r, column=1, value="  yoy %").font = FSM; ws.cell(row=r, column=1).border = BD
    for j in [1,2]:
        fval(ws, r, 2+j, f"={CL(2+j)}{r-1}/{CL(1+j)}{r-1}-1", NP); ws.cell(row=r, column=2+j).font = FSM
    for yr_j in range(5):
        pc=5+yr_j; fval(ws, r, pc, f"={CL(pc)}{r-1}/{CL(pc-1)}{r-1}-1", NP); ws.cell(row=r, column=pc).font = FSM
    r += 2

    # Google Services Total
    gs_total_r = r
    ws.cell(row=r, column=1, value="Google Services Total 谷歌服务合计").font = FB
    sb(ws, r, 1, r, 9, fill=KEY, font=FB)
    for c in range(2, 10):
        fval(ws, r, c, f"={CL(c)}{ad_total_r}+{CL(c)}{rev_rows[seg_names[3]]}")
    note(ws, r, "=广告合计+订阅/平台/设备")
    R["gs_total"] = gs_total_r
    r += 2

    # Cloud
    sec(ws, r, "Google Cloud 谷歌云"); r += 1
    name = seg_names[4]
    ws.cell(row=r, column=1, value=name).font = FB; ws.cell(row=r, column=1).border = BD
    rev_rows[name] = r
    for j, v in enumerate(SEG_REV[name]):
        hval(ws, r, 2+j, v)
    for yr_j in range(5):
        pc = 5+yr_j; prev = "D" if yr_j == 0 else CL(pc-1)
        fval(ws, r, pc, f"={prev}{r}*(1+{gf(4, yr_j)})")
    note(ws, r, seg_notes_rev[name])
    R["cloud_rev"] = r
    r += 1
    ws.cell(row=r, column=1, value="  yoy %").font = FSM; ws.cell(row=r, column=1).border = BD
    for j in [1,2]:
        fval(ws, r, 2+j, f"={CL(2+j)}{r-1}/{CL(1+j)}{r-1}-1", NP); ws.cell(row=r, column=2+j).font = FSM
    for yr_j in range(5):
        pc=5+yr_j; fval(ws, r, pc, f"={CL(pc)}{r-1}/{CL(pc-1)}{r-1}-1", NP); ws.cell(row=r, column=pc).font = FSM
    r += 2

    # Other Bets
    sec(ws, r, "Other Bets 其他创新业务"); r += 1
    name = seg_names[5]
    ws.cell(row=r, column=1, value=name).font = FB; ws.cell(row=r, column=1).border = BD
    rev_rows[name] = r
    for j, v in enumerate(SEG_REV[name]):
        hval(ws, r, 2+j, v)
    for yr_j in range(5):
        pc = 5+yr_j; prev = "D" if yr_j == 0 else CL(pc-1)
        fval(ws, r, pc, f"={prev}{r}*(1+{gf(5, yr_j)})")
    note(ws, r, seg_notes_rev[name])
    R["ob_rev"] = r
    r += 1
    ws.cell(row=r, column=1, value="  yoy %").font = FSM; ws.cell(row=r, column=1).border = BD
    for j in [1,2]:
        fval(ws, r, 2+j, f"={CL(2+j)}{r-1}/{CL(1+j)}{r-1}-1", NP); ws.cell(row=r, column=2+j).font = FSM
    for yr_j in range(5):
        pc=5+yr_j; fval(ws, r, pc, f"={CL(pc)}{r-1}/{CL(pc-1)}{r-1}-1", NP); ws.cell(row=r, column=pc).font = FSM
    r += 2

    # Hedging
    hedge_r = r
    ws.cell(row=r, column=1, value="Hedging gains (losses) 外汇对冲").font = FN; ws.cell(row=r, column=1).border = BD
    for j, v in enumerate(HEDGE):
        hval(ws, r, 2+j, v)
    for yr_j in range(5):
        fval(ws, r, 5+yr_j, 0)
    note(ws, r, "假设为零 Assume zero")
    r += 2

    # TOTAL
    total_rev_r = r
    ws.cell(row=r, column=1, value="Total Revenue 总收入").font = FB
    sb(ws, r, 1, r, 9, fill=KEY, font=FB)
    for c in range(2, 10):
        fval(ws, r, c, f"={CL(c)}{gs_total_r}+{CL(c)}{R['cloud_rev']}+{CL(c)}{R['ob_rev']}+{CL(c)}{hedge_r}")
    note(ws, r, "=Services+Cloud+OtherBets+Hedging")
    R["total_rev"] = total_rev_r
    r += 1
    ws.cell(row=r, column=1, value="  yoy %").font = FSM; ws.cell(row=r, column=1).border = BD
    for j in range(1, 8):
        fval(ws, r, 2+j, f"={CL(2+j)}{total_rev_r}/{CL(1+j)}{total_rev_r}-1", NP)
        ws.cell(row=r, column=2+j).font = FSM
    r += 2

    # Revenue Mix
    sec(ws, r, "Revenue Mix % 收入结构"); r += 1
    for name in ["Google Search & other", "YouTube ads", "Google Network",
                  "Subs, Platforms & Devices", "Google Cloud", "Other Bets"]:
        ws.cell(row=r, column=1, value=f"  {name}").font = FSM; ws.cell(row=r, column=1).border = BD
        for c in range(2, 10):
            fval(ws, r, c, f"={CL(c)}{rev_rows[name]}/{CL(c)}{total_rev_r}", NP)
            ws.cell(row=r, column=c).font = FSM
        r += 1
    R["ad_total"] = ad_total_r

    ws.freeze_panes = "B5"
    add_legend(ws, r + 2)

    # ═══════════════════════════════════════════════════
    #  TAB 5: SEGMENT P&L
    # ═══════════════════════════════════════════════════
    ws = wb.create_sheet("Segment_PL")
    ws["A1"] = "Segment P&L 分部利润表 (USD mm)"; ws["A1"].font = FT
    ws["A2"] = "Revenue / Costs / Operating Income by Division"; ws["A2"].font = FSM
    cw(ws, {CL(c): (44 if c == 1 else (38 if c == NCOL else 15)) for c in range(1, NCOL+1)})
    yh(ws, 3)

    r = 5
    # ── Google Services ──
    sec(ws, r, "Google Services 谷歌服务"); r += 1
    svc_rev_r = r
    ws.cell(row=r, column=1, value="Revenue 收入").font = FB; ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        lval(ws, r, c, f"=Segment_Revenue!{CL(c)}{gs_total_r}")
    note(ws, r, "链接自Segment_Revenue")
    r += 1
    svc_emp_r = r
    ws.cell(row=r, column=1, value="(-) Employee Comp 员工薪酬").font = FN; ws.cell(row=r, column=1).border = BD
    for j, v in enumerate(SEG_PL["svc_emp"]):
        hval(ws, r, 2+j, v)
    for c in range(5, 10):
        fval(ws, r, c, f"={CL(c)}{svc_rev_r}*Assumptions!$I$52")
    note(ws, r, "=Revenue×Emp Comp%, 含SBC+福利 (Note 15)")
    r += 1
    svc_oth_r = r
    ws.cell(row=r, column=1, value="(-) Other Costs 其他费用").font = FN; ws.cell(row=r, column=1).border = BD
    for j, v in enumerate(SEG_PL["svc_oth"]):
        hval(ws, r, 2+j, v)
    for c in range(5, 10):
        fval(ws, r, c, f"={CL(c)}{svc_rev_r}*Assumptions!$I$53")
    note(ws, r, "含TAC/内容采购/基础设施/设备成本")
    r += 1
    svc_tc_r = r
    ws.cell(row=r, column=1, value="Total Costs 总成本").font = FB; ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        fval(ws, r, c, f"={CL(c)}{svc_emp_r}+{CL(c)}{svc_oth_r}")
    r += 1
    svc_oi_r = r
    ws.cell(row=r, column=1, value="Operating Income 经营利润").font = FB; sb(ws, r, 1, r, 9, fill=KEY, font=FB)
    for c in range(2, 10):
        fval(ws, r, c, f"={CL(c)}{svc_rev_r}-{CL(c)}{svc_tc_r}")
    note(ws, r, "=Revenue-Total Costs")
    r += 1
    ws.cell(row=r, column=1, value="  Op Margin 经营利润率").font = FSM; ws.cell(row=r, column=1).border = BD
    pct_row(ws, r, svc_oi_r, svc_rev_r)
    r += 2

    # ── Google Cloud ──
    sec(ws, r, "Google Cloud 谷歌云"); r += 1
    cld_rev_r = r
    ws.cell(row=r, column=1, value="Revenue 收入").font = FB; ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        lval(ws, r, c, f"=Segment_Revenue!{CL(c)}{R['cloud_rev']}")
    note(ws, r, "链接自Segment_Revenue")
    r += 1
    cld_emp_r = r
    ws.cell(row=r, column=1, value="(-) Employee Comp 员工薪酬").font = FN; ws.cell(row=r, column=1).border = BD
    for j, v in enumerate(SEG_PL["cld_emp"]):
        hval(ws, r, 2+j, v)
    for c in range(5, 10):
        fval(ws, r, c, f"={CL(c)}{cld_rev_r}*Assumptions!$I$54")
    note(ws, r, "=Revenue×Cloud Emp%, 人效持续改善")
    r += 1
    cld_oth_r = r
    ws.cell(row=r, column=1, value="(-) Other Costs 其他费用").font = FN; ws.cell(row=r, column=1).border = BD
    for j, v in enumerate(SEG_PL["cld_oth"]):
        hval(ws, r, 2+j, v)
    for c in range(5, 10):
        fval(ws, r, c, f"={CL(c)}{cld_rev_r}*Assumptions!$I$55")
    note(ws, r, "含基础设施/第三方服务费")
    r += 1
    cld_tc_r = r
    ws.cell(row=r, column=1, value="Total Costs 总成本").font = FB; ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        fval(ws, r, c, f"={CL(c)}{cld_emp_r}+{CL(c)}{cld_oth_r}")
    r += 1
    cld_oi_r = r
    ws.cell(row=r, column=1, value="Operating Income 经营利润").font = FB; sb(ws, r, 1, r, 9, fill=KEY, font=FB)
    for c in range(2, 10):
        fval(ws, r, c, f"={CL(c)}{cld_rev_r}-{CL(c)}{cld_tc_r}")
    note(ws, r, "Cloud利润率快速提升")
    r += 1
    ws.cell(row=r, column=1, value="  Op Margin 经营利润率").font = FSM; ws.cell(row=r, column=1).border = BD
    pct_row(ws, r, cld_oi_r, cld_rev_r)
    r += 2

    # ── Other Bets ──
    sec(ws, r, "Other Bets 其他创新业务"); r += 1
    ob_rev_r2 = r
    ws.cell(row=r, column=1, value="Revenue 收入").font = FB; ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        lval(ws, r, c, f"=Segment_Revenue!{CL(c)}{R['ob_rev']}")
    r += 1
    ob_oi_r = r
    ws.cell(row=r, column=1, value="Operating Income (Loss) 经营利润(亏损)").font = FB; ws.cell(row=r, column=1).border = BD
    for j, v in enumerate(SEG_PL["ob_oi"]):
        hval(ws, r, 2+j, v)
    for c in range(5, 10):
        fval(ws, r, c, f"={CL(c)}{ob_rev_r2}*Assumptions!$I$56")
    note(ws, r, "=Revenue×OI Margin%, 前沿业务持续亏损")
    r += 1
    ws.cell(row=r, column=1, value="Implied Costs 隐含成本 (=Rev-OI)").font = FN; ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        fval(ws, r, c, f"={CL(c)}{ob_rev_r2}-{CL(c)}{ob_oi_r}")
    r += 1
    ws.cell(row=r, column=1, value="  Op Margin").font = FSM; ws.cell(row=r, column=1).border = BD
    pct_row(ws, r, ob_oi_r, ob_rev_r2)
    r += 2

    # ── Alphabet-level ──
    sec(ws, r, "Alphabet-level 未分配(AI研发+企业费用)"); r += 1
    alpha_r = r
    ws.cell(row=r, column=1, value="Unallocated Costs 未分配费用").font = FB; ws.cell(row=r, column=1).border = BD
    for j, v in enumerate(SEG_PL["alpha"]):
        hval(ws, r, 2+j, v)
    for c in range(5, 10):
        fval(ws, r, c, f"=-Assumptions!$I$57")
    note(ws, r, "未分配AI研发+企业管理费用")
    r += 1
    ws.cell(row=r, column=1, value="  % of Total Revenue").font = FSM; ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        fval(ws, r, c, f"={CL(c)}{alpha_r}/Segment_Revenue!{CL(c)}{total_rev_r}", NP)
        ws.cell(row=r, column=c).font = FSM
    r += 2

    # ── Cross-Check ──
    sec(ws, r, "Cross-Check 交叉验证: Sum of Segment OI vs Consolidated EBIT"); r += 1
    sum_seg_r = r
    R["sum_seg"] = sum_seg_r
    ws.cell(row=r, column=1, value="Sum of Segment OI 分部OI汇总").font = FB; ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        fval(ws, r, c, f"={CL(c)}{svc_oi_r}+{CL(c)}{cld_oi_r}+{CL(c)}{ob_oi_r}+{CL(c)}{alpha_r}")
    note(ws, r, "=Services+Cloud+OtherBets+Alphabet-level")
    r += 1
    consol_ebit_link_r = r
    R["consol_ebit_link"] = r
    ws.cell(row=r, column=1, value="Consolidated EBIT (from Consolidated_PL)").font = FB; ws.cell(row=r, column=1).border = BD
    # Fill after building Consolidated_PL
    r += 1
    diff_r = r
    ws.cell(row=r, column=1, value="Difference 差异 (should = 0)").font = FB; ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        fval(ws, r, c, f"={CL(c)}{sum_seg_r}-{CL(c)}{consol_ebit_link_r}")
    sb(ws, diff_r, 1, diff_r, 9, fill=ERR, font=FB)
    note(ws, diff_r, "应为0，否则分部与合并口径不一致")

    ws.freeze_panes = "B4"
    add_legend(ws, r + 3)

    # ═══════════════════════════════════════════════════
    #  TAB 6: CONSOLIDATED P&L
    # ═══════════════════════════════════════════════════
    ws = wb.create_sheet("Consolidated_PL")
    ws["A1"] = "Consolidated Income Statement 合并利润表 (USD mm)"; ws["A1"].font = FT
    ws["A2"] = "Source: 10-K p.49"; ws["A2"].font = FSM
    cw(ws, {CL(c): (44 if c == 1 else (38 if c == NCOL else 15)) for c in range(1, NCOL+1)})
    yh(ws, 3)

    r = 5
    rev_pl = r
    ws.cell(row=r, column=1, value="Total Revenue 总收入").font = FB; sb(ws, r, 1, r, 9, fill=KEY, font=FB)
    for c in range(2, 10):
        lval(ws, r, c, f"=Segment_Revenue!{CL(c)}{total_rev_r}")
    note(ws, r, "链接自Segment_Revenue")
    R["rev_pl"] = rev_pl
    r += 1
    ws.cell(row=r, column=1, value="  yoy %").font = FSM; ws.cell(row=r, column=1).border = BD
    for j in range(1, 8):
        fval(ws, r, 2+j, f"={CL(2+j)}{rev_pl}/{CL(1+j)}{rev_pl}-1", NP); ws.cell(row=r, column=2+j).font = FSM
    r += 1

    ad_pl = r
    ws.cell(row=r, column=1, value="  Google Advertising Rev (memo) 广告收入").font = FN; ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        lval(ws, r, c, f"=Segment_Revenue!{CL(c)}{ad_total_r}")
    note(ws, r, "用于计算TAC")
    r += 2

    # COGS
    sec(ws, r, "Cost of Revenues 营业成本 (COGS)"); r += 1
    tac_pl = r
    ws.cell(row=r, column=1, value="  TAC 流量获取成本").font = FN; ws.cell(row=r, column=1).border = BD
    for j, v in enumerate(PL["TAC"]):
        hval(ws, r, 2+j, v)
    for c in range(5, 10):
        fval(ws, r, c, f"={CL(c)}{ad_pl}*Assumptions!$I$61")
    note(ws, r, "=Ad Revenue×TAC Rate")
    r += 1
    ws.cell(row=r, column=1, value="    TAC rate (% of Ad Rev)").font = FSM; ws.cell(row=r, column=1).border = BD
    pct_row(ws, r, tac_pl, ad_pl); r += 1

    ocogs_pl = r
    ws.cell(row=r, column=1, value="  Other Cost of Revenues 其他营业成本").font = FN; ws.cell(row=r, column=1).border = BD
    for j, v in enumerate(PL["OtherCOGS"]):
        hval(ws, r, 2+j, v)
    for c in range(5, 10):
        fval(ws, r, c, f"={CL(c)}{rev_pl}*Assumptions!$I$62")
    note(ws, r, "基础设施折旧/内容/数据中心运营")
    r += 1
    ws.cell(row=r, column=1, value="    Other COGS %").font = FSM; ws.cell(row=r, column=1).border = BD
    pct_row(ws, r, ocogs_pl, rev_pl); r += 1

    cogs_pl = r
    ws.cell(row=r, column=1, value="Total COGS 营业成本合计").font = FB; ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        fval(ws, r, c, f"={CL(c)}{tac_pl}+{CL(c)}{ocogs_pl}")
    R["cogs_pl"] = cogs_pl
    r += 1

    gp_pl = r
    ws.cell(row=r, column=1, value="Gross Profit 毛利润").font = FB; sb(ws, r, 1, r, 9, fill=SUB, font=FB)
    for c in range(2, 10):
        fval(ws, r, c, f"={CL(c)}{rev_pl}-{CL(c)}{cogs_pl}")
    note(ws, r, "=Revenue-COGS")
    r += 1
    ws.cell(row=r, column=1, value="  Gross Margin 毛利率").font = FSM; ws.cell(row=r, column=1).border = BD
    pct_row(ws, r, gp_pl, rev_pl)
    r += 2

    # OpEx
    sec(ws, r, "Operating Expenses 营业费用 (OpEx)"); r += 1
    opex_spec = [
        ("  R&D 研发费用", PL["R&D"], 63, "AI+Cloud持续投入，规模效应显现"),
        ("  S&M 销售与营销费用", PL["S&M"], 64, "效率提升，费率持续下降"),
        ("  G&A 管理费用 (含一次性项目)", PL["G&A"], 65, "2025含DOJ诉讼拨备，正常化后下降"),
    ]
    opex_rows = []
    for name, hist, ref_row, nt_text in opex_spec:
        ws.cell(row=r, column=1, value=name).font = FN; ws.cell(row=r, column=1).border = BD
        for j, v in enumerate(hist):
            hval(ws, r, 2+j, v)
        for c in range(5, 10):
            fval(ws, r, c, f"={CL(c)}{rev_pl}*Assumptions!$I${ref_row}")
        note(ws, r, nt_text)
        opex_rows.append(r); r += 1
        ws.cell(row=r, column=1, value=f"    {name.strip().split('(')[0].strip()} %").font = FSM; ws.cell(row=r, column=1).border = BD
        pct_row(ws, r, r-1, rev_pl); r += 1

    topex_pl = r
    ws.cell(row=r, column=1, value="Estimated Total OpEx (R&D+S&M+G&A) 营业费用合计").font = FB
    ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        fval(ws, r, c, "=" + "+".join(f"{CL(c)}{x}" for x in opex_rows))
    note(ws, r, "参考性合计，EBIT由分部OI驱动")
    r += 1

    # EBIT = segment-driven
    ebit_pl = r
    ws.cell(row=r, column=1, value="EBIT 经营利润 (=Sum of Segment OI)").font = FB
    sb(ws, r, 1, r, 9, fill=KEY, font=FB)
    for c in range(2, 10):
        lval(ws, r, c, f"=Segment_PL!{CL(c)}{sum_seg_r}")
    note(ws, r, "分部驱动：链接自Segment_PL汇总OI")
    R["ebit_pl"] = ebit_pl
    r += 1
    ws.cell(row=r, column=1, value="  EBIT Margin 经营利润率").font = FSM; ws.cell(row=r, column=1).border = BD
    pct_row(ws, r, ebit_pl, rev_pl)
    r += 1
    ws.cell(row=r, column=1, value="  Implied OpEx (=GP-EBIT) 隐含OpEx").font = FSM; ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        fval(ws, r, c, f"={CL(c)}{gp_pl}-{CL(c)}{ebit_pl}")
        ws.cell(row=r, column=c).font = FSM
    r += 2

    # OI&E
    oie_pl = r
    ws.cell(row=r, column=1, value="OI&E 其他收入/(支出)").font = FN; ws.cell(row=r, column=1).border = BD
    for j, v in enumerate(PL["OIE"]):
        hval(ws, r, 2+j, v)
    for j, v in enumerate([5000, 5500, 6000, 6500, 7000]):
        c = ws.cell(row=r, column=5+j, value=v); c.number_format = NI; c.border = BD; c.fill = BLUE; c.font = FI
    note(ws, r, "2025含$24B非上市证券估值收益(一次性)")
    r += 1

    pbt_pl = r
    ws.cell(row=r, column=1, value="Pre-tax Income 税前利润").font = FB; ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        fval(ws, r, c, f"={CL(c)}{ebit_pl}+{CL(c)}{oie_pl}")
    r += 1
    tax_pl = r
    ws.cell(row=r, column=1, value="Income Tax 所得税").font = FN; ws.cell(row=r, column=1).border = BD
    for j, v in enumerate(PL["Tax"]):
        hval(ws, r, 2+j, v)
    for c in range(5, 10):
        fval(ws, r, c, f"={CL(c)}{pbt_pl}*Assumptions!$I$11")
    note(ws, r, "=Pre-tax Income×ETR")
    r += 1

    ni_pl = r
    ws.cell(row=r, column=1, value="Net Income 净利润").font = FB; sb(ws, r, 1, r, 9, fill=KEY, font=FB)
    for c in range(2, 10):
        fval(ws, r, c, f"={CL(c)}{pbt_pl}-{CL(c)}{tax_pl}")
    note(ws, r, "=Pre-tax - Tax")
    R["ni_pl"] = ni_pl
    r += 1
    ws.cell(row=r, column=1, value="  Net Margin 净利率").font = FSM; ws.cell(row=r, column=1).border = BD
    pct_row(ws, r, ni_pl, rev_pl)
    r += 1
    ws.cell(row=r, column=1, value="  NI yoy %").font = FSM; ws.cell(row=r, column=1).border = BD
    for j in range(1, 8):
        fval(ws, r, 2+j, f"={CL(2+j)}{ni_pl}/{CL(1+j)}{ni_pl}-1", NP); ws.cell(row=r, column=2+j).font = FSM
    r += 2

    # Memo items
    da_pl = r
    ws.cell(row=r, column=1, value="D&A 折旧与摊销 (memo)").font = FB; ws.cell(row=r, column=1).border = BD
    for j, v in enumerate(PL["D&A"]):
        hval(ws, r, 2+j, v)
    for c in range(5, 10):
        fval(ws, r, c, f"={CL(c)}{rev_pl}*Assumptions!$I$66")
    note(ws, r, "CapEx高增长→D&A占比上升")
    R["da_pl"] = da_pl
    r += 1

    sbc_pl = r
    ws.cell(row=r, column=1, value="SBC 股权激励 (memo)").font = FB; ws.cell(row=r, column=1).border = BD
    for j, v in enumerate(PL["SBC"]):
        hval(ws, r, 2+j, v)
    for c in range(5, 10):
        fval(ws, r, c, f"={CL(c)}{rev_pl}*Assumptions!$I$67")
    note(ws, r, "SBC占比逐步下降")
    R["sbc_pl"] = sbc_pl
    r += 2

    # EPS
    shares_pl = r
    ws.cell(row=r, column=1, value="Diluted Shares 稀释股数 (mm)").font = FB; ws.cell(row=r, column=1).border = BD
    for j, v in enumerate(SHARES):
        hval(ws, r, 2+j, v)
    for yr_j in range(5):
        pc = 5+yr_j; prev = "D" if yr_j == 0 else CL(pc-1)
        fval(ws, r, pc, f"={prev}{r}*0.98")
    note(ws, r, "年回购~2%股本")
    R["shares_pl"] = shares_pl
    r += 1

    eps_pl = r
    ws.cell(row=r, column=1, value="Diluted EPS 稀释每股收益").font = FB; sb(ws, r, 1, r, 9, fill=KEY, font=FB)
    for c in range(2, 10):
        fval(ws, r, c, f"={CL(c)}{ni_pl}/{CL(c)}{shares_pl}", NU)
    note(ws, r, "=Net Income / Diluted Shares")
    R["eps_pl"] = eps_pl

    ws.freeze_panes = "B4"
    add_legend(ws, r + 3)

    # ── Fill Segment_PL cross-check link ──
    ws_seg = wb["Segment_PL"]
    for c in range(2, 10):
        lval(ws_seg, R["consol_ebit_link"], c, f"=Consolidated_PL!{CL(c)}{ebit_pl}")

    # ═══════════════════════════════════════════════════
    #  TAB 7: BALANCE SHEET (NEW)
    # ═══════════════════════════════════════════════════
    ws = wb.create_sheet("BS")
    ws["A1"] = "Consolidated Balance Sheet 合并资产负债表 (USD mm)"; ws["A1"].font = FT
    ws["A2"] = "Source: 10-K p.48 | 2023 estimated from prior filing"; ws["A2"].font = FSM
    cw(ws, {CL(c): (44 if c == 1 else (38 if c == NCOL else 15)) for c in range(1, NCOL+1)})
    yh(ws, 3)

    r = 5
    sec(ws, r, "Current Assets 流动资产"); r += 1
    bs_items_ca = [
        ("Cash & Equivalents 现金", "Cash", "链接自CF期末现金(forecast)"),
        ("Marketable Securities 有价证券", "MktSec", "投资组合，假设温和增长"),
        ("Total Cash & Securities 现金合计", "CashTotal", "=Cash+Securities"),
        ("Accounts Receivable, net 应收账款", "AR", "=Revenue×AR Days/365"),
        ("Other Current Assets 其他流动资产", "OtherCA", "预付/库存/其他"),
    ]
    bs_rows = {}
    for label, key, nt in bs_items_ca:
        ws.cell(row=r, column=1, value=label).font = FN if "Total" not in label else FB
        ws.cell(row=r, column=1).border = BD
        for j, v in enumerate(BS_DATA[key]):
            hval(ws, r, 2+j, v)
        bs_rows[key] = r
        note(ws, r, nt)
        r += 1

    # Forecast current assets
    # AR = Revenue × AR Days / 365
    for c in range(5, 10):
        fval(ws, bs_rows["AR"], c, f"=Consolidated_PL!{CL(c)}{rev_pl}*Assumptions!$I$84/365")
    # Other current: ~4% of revenue
    for c in range(5, 10):
        fval(ws, bs_rows["OtherCA"], c, f"=Consolidated_PL!{CL(c)}{rev_pl}*0.04")
    # Marketable securities: grow 5% per year
    for c in range(5, 10):
        prev = "D" if c == 5 else CL(c-1)
        fval(ws, bs_rows["MktSec"], c, f"={prev}{bs_rows['MktSec']}*1.05")
    # Cash: will be filled as plug later
    # Cash Total
    for c in range(5, 10):
        fval(ws, bs_rows["CashTotal"], c, f"={CL(c)}{bs_rows['Cash']}+{CL(c)}{bs_rows['MktSec']}")

    total_ca_r = r
    ws.cell(row=r, column=1, value="Total Current Assets 流动资产合计").font = FB
    sb(ws, r, 1, r, 9, fill=SUB, font=FB)
    for c in range(2, 10):
        fval(ws, r, c, f"={CL(c)}{bs_rows['CashTotal']}+{CL(c)}{bs_rows['AR']}+{CL(c)}{bs_rows['OtherCA']}")
    r += 2

    sec(ws, r, "Non-Current Assets 非流动资产"); r += 1
    bs_items_nca = [
        ("Non-marketable Securities 非上市证券", "NonMktSec", "含私募股权投资"),
        ("Deferred Income Taxes 递延所得税", "DeferredTax", ""),
        ("PP&E, net 固定资产净值", "PPE", "=Prior+CapEx-D&A"),
        ("Operating Lease Assets 经营租赁", "OpLeaseAsset", "办公/数据中心租赁"),
        ("Goodwill 商誉", "Goodwill", "收购溢价"),
        ("Other Non-current 其他非流动", "OtherNCA", "无形资产/其他"),
    ]
    for label, key, nt in bs_items_nca:
        ws.cell(row=r, column=1, value=label).font = FN; ws.cell(row=r, column=1).border = BD
        for j, v in enumerate(BS_DATA[key]):
            hval(ws, r, 2+j, v)
        bs_rows[key] = r
        note(ws, r, nt)
        r += 1

    # Forecast non-current
    # PP&E = prior + CapEx - D&A
    for yr_j in range(5):
        c = 5 + yr_j; prev = "D" if yr_j == 0 else CL(c-1)
        fval(ws, bs_rows["PPE"], c, f"={prev}{bs_rows['PPE']}+Assumptions!$I${capex_rows[yr_j]}-Consolidated_PL!{CL(c)}{da_pl}")
    # Non-marketable: grow per assumption
    for c in range(5, 10):
        prev = "D" if c == 5 else CL(c-1)
        fval(ws, bs_rows["NonMktSec"], c, f"={prev}{bs_rows['NonMktSec']}*(1+Assumptions!$I$91)")
    # Deferred tax: assume stable ~$9B
    for c in range(5, 10):
        fval(ws, bs_rows["DeferredTax"], c, 9000)
    # Operating lease: grow per assumption
    for c in range(5, 10):
        prev = "D" if c == 5 else CL(c-1)
        fval(ws, bs_rows["OpLeaseAsset"], c, f"={prev}{bs_rows['OpLeaseAsset']}*(1+Assumptions!$I$90)")
    # Goodwill: stable
    for c in range(5, 10):
        prev = "D" if c == 5 else CL(c-1)
        fval(ws, bs_rows["Goodwill"], c, f"={prev}{bs_rows['Goodwill']}*1.02")
    # Other NCA: stable growth
    for c in range(5, 10):
        prev = "D" if c == 5 else CL(c-1)
        fval(ws, bs_rows["OtherNCA"], c, f"={prev}{bs_rows['OtherNCA']}*1.03")

    total_nca_r = r
    ws.cell(row=r, column=1, value="Total Non-Current Assets 非流动资产合计").font = FB
    sb(ws, r, 1, r, 9, fill=SUB, font=FB)
    nca_items = ["NonMktSec", "DeferredTax", "PPE", "OpLeaseAsset", "Goodwill", "OtherNCA"]
    for c in range(2, 10):
        fval(ws, r, c, "=" + "+".join(f"{CL(c)}{bs_rows[k]}" for k in nca_items))
    r += 1

    total_assets_r = r
    ws.cell(row=r, column=1, value="TOTAL ASSETS 资产合计").font = FB; sb(ws, r, 1, r, 9, fill=KEY, font=FB)
    for c in range(2, 10):
        fval(ws, r, c, f"={CL(c)}{total_ca_r}+{CL(c)}{total_nca_r}")
    R["total_assets"] = total_assets_r
    r += 2

    # LIABILITIES
    sec(ws, r, "Current Liabilities 流动负债"); r += 1
    bs_items_cl = [
        ("Accounts Payable 应付账款", "AP", "=COGS×AP Days/365"),
        ("Accrued Compensation 应计薪酬", "AccComp", "=Revenue×AccComp%"),
        ("Accrued Expenses 应计费用", "AccExp", "=Revenue×AccExp%"),
        ("Accrued Revenue Share 应计收入分成", "RevShare", "=Ad Revenue×RevShare%"),
        ("Deferred Revenue 递延收入", "DefRev", "Cloud预付合同增长"),
    ]
    for label, key, nt in bs_items_cl:
        ws.cell(row=r, column=1, value=label).font = FN; ws.cell(row=r, column=1).border = BD
        for j, v in enumerate(BS_DATA[key]):
            hval(ws, r, 2+j, v)
        bs_rows[key] = r
        note(ws, r, nt)
        r += 1

    # Forecast CL
    for c in range(5, 10):
        fval(ws, bs_rows["AP"], c, f"=Consolidated_PL!{CL(c)}{R['cogs_pl']}*Assumptions!$I$85/365")
        fval(ws, bs_rows["AccComp"], c, f"=Consolidated_PL!{CL(c)}{rev_pl}*Assumptions!$I$86")
        fval(ws, bs_rows["AccExp"], c, f"=Consolidated_PL!{CL(c)}{rev_pl}*Assumptions!$I$87")
        fval(ws, bs_rows["RevShare"], c, f"=Segment_Revenue!{CL(c)}{ad_total_r}*Assumptions!$I$88")
    for c in range(5, 10):
        prev = "D" if c == 5 else CL(c-1)
        fval(ws, bs_rows["DefRev"], c, f"={prev}{bs_rows['DefRev']}*(1+Assumptions!$I$89)")

    total_cl_r = r
    ws.cell(row=r, column=1, value="Total Current Liabilities 流动负债合计").font = FB
    sb(ws, r, 1, r, 9, fill=SUB, font=FB)
    cl_items = ["AP", "AccComp", "AccExp", "RevShare", "DefRev"]
    for c in range(2, 10):
        fval(ws, r, c, "=" + "+".join(f"{CL(c)}{bs_rows[k]}" for k in cl_items))
    r += 2

    sec(ws, r, "Non-Current Liabilities 非流动负债"); r += 1
    bs_items_ncl = [
        ("Long-term Debt 长期债务", "LTDebt", "2025发行$65B新债"),
        ("Income Taxes Payable, NC 长期应交税", "TaxNC", ""),
        ("Operating Lease Liabilities 经营租赁负债", "OpLeaseLiab", "跟踪ROU资产"),
        ("Other Long-term 其他长期负债", "OtherLTL", ""),
    ]
    for label, key, nt in bs_items_ncl:
        ws.cell(row=r, column=1, value=label).font = FN; ws.cell(row=r, column=1).border = BD
        for j, v in enumerate(BS_DATA[key]):
            hval(ws, r, 2+j, v)
        bs_rows[key] = r
        note(ws, r, nt)
        r += 1

    # Forecast NCL
    for c in range(5, 10):
        fval(ws, bs_rows["LTDebt"], c, f"=Assumptions!$I$22")  # from assumption
        prev = "D" if c == 5 else CL(c-1)
        fval(ws, bs_rows["TaxNC"], c, f"={prev}{bs_rows['TaxNC']}*1.03")
        fval(ws, bs_rows["OpLeaseLiab"], c, f"={CL(c)}{bs_rows['OpLeaseAsset']}*0.84")  # ~84% of ROU
        fval(ws, bs_rows["OtherLTL"], c, f"={prev}{bs_rows['OtherLTL']}*1.05")

    total_ncl_r = r
    ncl_items = ["LTDebt", "TaxNC", "OpLeaseLiab", "OtherLTL"]
    ws.cell(row=r, column=1, value="Total Non-Current Liabilities 非流动负债").font = FB
    ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        fval(ws, r, c, "=" + "+".join(f"{CL(c)}{bs_rows[k]}" for k in ncl_items))
    r += 1

    total_liab_r = r
    ws.cell(row=r, column=1, value="TOTAL LIABILITIES 负债合计").font = FB; sb(ws, r, 1, r, 9, fill=SUB, font=FB)
    for c in range(2, 10):
        fval(ws, r, c, f"={CL(c)}{total_cl_r}+{CL(c)}{total_ncl_r}")
    r += 2

    # EQUITY
    sec(ws, r, "Stockholders' Equity 股东权益"); r += 1
    equity_r = r
    ws.cell(row=r, column=1, value="Total Equity 股东权益合计").font = FB; sb(ws, r, 1, r, 9, fill=KEY, font=FB)
    for j, v in enumerate(BS_DATA["Equity"]):
        hval(ws, r, 2+j, v)
    # Equity forecast: prior + NI - Buyback - Dividends + SBC (net of tax withholding)
    for c in range(5, 10):
        prev = "D" if c == 5 else CL(c-1)
        fval(ws, r, c,
             f"={prev}{equity_r}+Consolidated_PL!{CL(c)}{ni_pl}"
             f"-Assumptions!$I$79"
             f"-Assumptions!$I$80*Consolidated_PL!{CL(c)}{R['shares_pl']}"
             f"+Consolidated_PL!{CL(c)}{sbc_pl}*0.5")  # net of tax withholding ~50%
    note(ws, r, "=Prior+NI-Buyback-Div+SBC(net)")
    R["equity"] = equity_r
    r += 2

    # Total L&E
    total_le_r = r
    ws.cell(row=r, column=1, value="TOTAL LIABILITIES & EQUITY 负债+权益").font = FB
    sb(ws, r, 1, r, 9, fill=KEY, font=FB)
    for c in range(2, 10):
        fval(ws, r, c, f"={CL(c)}{total_liab_r}+{CL(c)}{equity_r}")
    r += 1

    # Balance check
    chk_r = r
    ws.cell(row=r, column=1, value="Balance Check 平衡检验 (Assets - L&E)").font = FB; ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        fval(ws, r, c, f"={CL(c)}{total_assets_r}-{CL(c)}{total_le_r}")
    sb(ws, chk_r, 1, chk_r, 9, fill=ERR, font=FB)
    note(ws, chk_r, "应为0 (Cash为plug项)")

    # Now fill Cash as plug: Cash = Total L&E - (all other assets)
    # Actually simpler: Cash = Total Equity + Total Liabilities - NCA - (AR + OtherCA + MktSec)
    for c in range(5, 10):
        fval(ws, bs_rows["Cash"], c,
             f"={CL(c)}{equity_r}+{CL(c)}{total_liab_r}-{CL(c)}{total_nca_r}"
             f"-{CL(c)}{bs_rows['MktSec']}-{CL(c)}{bs_rows['AR']}-{CL(c)}{bs_rows['OtherCA']}")

    ws.freeze_panes = "B4"
    add_legend(ws, r + 3)

    # ═══════════════════════════════════════════════════
    #  TAB 8: CASH FLOW (Enhanced with WC detail)
    # ═══════════════════════════════════════════════════
    ws = wb.create_sheet("Cash_Flow")
    ws["A1"] = "Cash Flow Statement 现金流量表 (USD mm)"; ws["A1"].font = FT
    ws["A2"] = "Source: 10-K pp.51-52"; ws["A2"].font = FSM
    cw(ws, {CL(c): (44 if c == 1 else (38 if c == NCOL else 15)) for c in range(1, NCOL+1)})
    yh(ws, 3)

    r = 5
    sec(ws, r, "Operating Activities 经营活动现金流"); r += 1

    cf_items_op = [
        ("Net Income 净利润", CF["NI"], f"Consolidated_PL!{{c}}{ni_pl}", "链接自Consolidated_PL"),
        ("(+) D&A 折旧与摊销", CF["DA"], f"Consolidated_PL!{{c}}{da_pl}", "非现金项目加回"),
        ("(+) SBC 股权激励", CF["SBC"], f"Consolidated_PL!{{c}}{sbc_pl}", "非现金项目加回"),
        ("Deferred Tax 递延所得税", CF["DeferredTax"], None, "递延税项变动"),
        ("(Gain)/Loss on Securities 证券损益", CF["SecGL"], None, "非现金投资损益"),
        ("Other Adjustments 其他调整", CF["Other"], None, "其他非现金项目"),
    ]
    cf_rows = {}
    for name, hist, tmpl, nt in cf_items_op:
        ws.cell(row=r, column=1, value=name).font = FN; ws.cell(row=r, column=1).border = BD
        for j, v in enumerate(hist):
            hval(ws, r, 2+j, v)
        for c in range(5, 10):
            if tmpl:
                lval(ws, r, c, f"={tmpl.format(c=CL(c))}")
            else:
                fval(ws, r, c, 0)
        cf_rows[name] = r
        note(ws, r, nt)
        r += 1

    # Working Capital changes
    sec(ws, r, "Working Capital Changes 营运资金变动"); r += 1
    wc_items = [
        ("  Chg in AR 应收变动", CF["ChgAR"], "ChgAR", "应收增加为负"),
        ("  Chg in Tax 应交税变动", CF["ChgTaxNet"], "ChgTax", ""),
        ("  Chg in Other Assets 其他资产变动", CF["ChgOtherA"], "ChgOA", ""),
        ("  Chg in AP 应付变动", CF["ChgAP"], "ChgAP", "应付增加为正"),
        ("  Chg in Accrued Expenses 应计费用变动", CF["ChgAccExp"], "ChgAcc", ""),
        ("  Chg in Revenue Share 收入分成变动", CF["ChgRevShare"], "ChgRS", ""),
        ("  Chg in Deferred Revenue 递延收入变动", CF["ChgDefRev"], "ChgDR", ""),
    ]
    wc_rows = []
    for name, hist, key, nt in wc_items:
        ws.cell(row=r, column=1, value=name).font = FN; ws.cell(row=r, column=1).border = BD
        for j, v in enumerate(hist):
            hval(ws, r, 2+j, v)
        # For forecast: link to BS changes
        for c in range(5, 10):
            fval(ws, r, c, 0)  # placeholder, we'll set formulas below
        wc_rows.append(r)
        cf_rows[key] = r
        note(ws, r, nt)
        r += 1

    # Set WC formulas linking to BS changes
    # ChgAR: -(current AR - prior AR)
    for c in range(5, 10):
        prev = "D" if c == 5 else CL(c-1)
        fval(ws, cf_rows["ChgAR"], c, f"=-(BS!{CL(c)}{bs_rows['AR']}-BS!{prev}{bs_rows['AR']})")
    # ChgAP: current AP - prior AP
    for c in range(5, 10):
        prev = "D" if c == 5 else CL(c-1)
        fval(ws, cf_rows["ChgAP"], c, f"=BS!{CL(c)}{bs_rows['AP']}-BS!{prev}{bs_rows['AP']}")
    # ChgAcc: current - prior
    for c in range(5, 10):
        prev = "D" if c == 5 else CL(c-1)
        fval(ws, cf_rows["ChgAcc"], c, f"=BS!{CL(c)}{bs_rows['AccExp']}-BS!{prev}{bs_rows['AccExp']}")
    # ChgRS:
    for c in range(5, 10):
        prev = "D" if c == 5 else CL(c-1)
        fval(ws, cf_rows["ChgRS"], c, f"=BS!{CL(c)}{bs_rows['RevShare']}-BS!{prev}{bs_rows['RevShare']}")
    # ChgDR:
    for c in range(5, 10):
        prev = "D" if c == 5 else CL(c-1)
        fval(ws, cf_rows["ChgDR"], c, f"=BS!{CL(c)}{bs_rows['DefRev']}-BS!{prev}{bs_rows['DefRev']}")

    cfo_r = r
    ws.cell(row=r, column=1, value="CFO 经营活动净现金").font = FB; sb(ws, r, 1, r, 9, fill=SUB, font=FB)
    for j, v in enumerate(CF["CFO"]):
        hval(ws, r, 2+j, v)
    all_cf_op = list(cf_rows.values())
    for c in range(5, 10):
        fval(ws, r, c, "=" + "+".join(f"{CL(c)}{x}" for x in all_cf_op))
    note(ws, r, "=NI+D&A+SBC+调整+WC变动")
    R["cfo"] = cfo_r
    r += 2

    # Investing
    sec(ws, r, "Investing Activities 投资活动"); r += 1
    capex_cf = r
    ws.cell(row=r, column=1, value="CapEx 资本支出").font = FN; ws.cell(row=r, column=1).border = BD
    for j, v in enumerate(CF["CapEx"]):
        hval(ws, r, 2+j, v)
    for yr_j in range(5):
        fval(ws, r, 5+yr_j, f"=-Assumptions!$I${capex_rows[yr_j]}")
    note(ws, r, "2026=$180B per mgmt guidance")
    R["capex_cf"] = capex_cf
    r += 1
    ws.cell(row=r, column=1, value="  CapEx / Revenue").font = FSM; ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        fval(ws, r, c, f"=-{CL(c)}{capex_cf}/Consolidated_PL!{CL(c)}{rev_pl}", NP)
        ws.cell(row=r, column=c).font = FSM
    r += 2

    fcf_r = r
    ws.cell(row=r, column=1, value="FCF 自由现金流 (=CFO+CapEx)").font = FB; sb(ws, r, 1, r, 9, fill=KEY, font=FB)
    for c in range(2, 10):
        fval(ws, r, c, f"={CL(c)}{cfo_r}+{CL(c)}{capex_cf}")
    note(ws, r, "=CFO-|CapEx|")
    R["fcf"] = fcf_r
    r += 1
    ws.cell(row=r, column=1, value="  FCF Margin 自由现金流利率").font = FSM; ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        fval(ws, r, c, f"={CL(c)}{fcf_r}/Consolidated_PL!{CL(c)}{rev_pl}", NP)
        ws.cell(row=r, column=c).font = FSM
    r += 2

    # Financing
    sec(ws, r, "Financing Activities 融资活动"); r += 1
    bb_r = r
    ws.cell(row=r, column=1, value="Share Repurchases 股份回购").font = FN; ws.cell(row=r, column=1).border = BD
    for j, v in enumerate(CF["Buyback"]):
        hval(ws, r, 2+j, v)
    for c in range(5, 10):
        fval(ws, r, c, "=-Assumptions!$I$79")
    note(ws, r, "链接自Assumptions回购假设")
    r += 1
    div_cf = r
    ws.cell(row=r, column=1, value="Dividends Paid 股息支付").font = FN; ws.cell(row=r, column=1).border = BD
    for j, v in enumerate(CF["Dividend"]):
        hval(ws, r, 2+j, v)
    for c in range(5, 10):
        fval(ws, r, c, f"=-Assumptions!$I$80*Consolidated_PL!{CL(c)}{R['shares_pl']}")
    note(ws, r, "=DPS×Diluted Shares")

    ws.freeze_panes = "B4"
    add_legend(ws, r + 3)

    # ═══════════════════════════════════════════════════
    #  TAB 9: DCF
    # ═══════════════════════════════════════════════════
    ws = wb.create_sheet("DCF")
    ws["A1"] = "DCF Valuation DCF估值 (USD mm)"; ws["A1"].font = FT
    ws["A2"] = "CapEx 2026 = $180B per mgmt guidance"; ws["A2"].font = FSM
    cw(ws, {"A": 44, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 38})
    for i, yr in enumerate(["2026E", "2027E", "2028E", "2029E", "2030E"]):
        c = ws.cell(row=4, column=2+i, value=yr); c.font = FW; c.fill = HDR; c.alignment = CT; c.border = BD
    ws.cell(row=4, column=1).fill = HDR; ws.cell(row=4, column=1).border = BD
    nc = ws.cell(row=4, column=7, value="Notes / 预测逻辑"); nc.font = FW; nc.fill = HDR; nc.alignment = CT; nc.border = BD

    r = 6
    ebit_dcf = r
    ws.cell(row=r, column=1, value="EBIT 经营利润").font = FB; ws.cell(row=r, column=1).border = BD
    for i in range(5):
        lval(ws, r, 2+i, f"=Consolidated_PL!{CL(5+i)}{ebit_pl}")
    ws.cell(row=r, column=7, value="链接自Consolidated_PL").font = FNOTE; ws.cell(row=r, column=7).border = BD
    r += 1
    tax_dcf = r
    ws.cell(row=r, column=1, value="(-) Tax on EBIT EBIT税负").font = FN; ws.cell(row=r, column=1).border = BD
    for i in range(5):
        fval(ws, r, 2+i, f"={CL(2+i)}{ebit_dcf}*Assumptions!$I$11")
    r += 1
    nopat_dcf = r
    ws.cell(row=r, column=1, value="NOPAT 税后净营业利润").font = FB; ws.cell(row=r, column=1).border = BD
    for i in range(5):
        fval(ws, r, 2+i, f"={CL(2+i)}{ebit_dcf}-{CL(2+i)}{tax_dcf}")
    ws.cell(row=r, column=7, value="EBIT×(1-Tax Rate)").font = FNOTE; ws.cell(row=r, column=7).border = BD
    r += 1
    da_dcf = r
    ws.cell(row=r, column=1, value="(+) D&A 折旧摊销").font = FN; ws.cell(row=r, column=1).border = BD
    for i in range(5):
        lval(ws, r, 2+i, f"=Consolidated_PL!{CL(5+i)}{da_pl}")
    r += 1
    capex_dcf = r
    ws.cell(row=r, column=1, value="(-) CapEx 资本支出").font = FN; ws.cell(row=r, column=1).border = BD
    for i in range(5):
        fval(ws, r, 2+i, f"=Assumptions!$I${capex_rows[i]}")
    ws.cell(row=r, column=7, value="2026=$180B management guidance").font = FNOTE; ws.cell(row=r, column=7).border = BD
    r += 1
    nwc_dcf = r
    ws.cell(row=r, column=1, value="(-) Change in NWC 净营运资本变动").font = FN; ws.cell(row=r, column=1).border = BD
    for i in range(5):
        src = CL(5+i); prev = "D" if i == 0 else CL(4+i)
        fval(ws, r, 2+i, f"=(Consolidated_PL!{src}{rev_pl}-Consolidated_PL!{prev}{rev_pl})*Assumptions!$I$68")
    r += 1

    ufcf_dcf = r
    ws.cell(row=r, column=1, value="UFCF 无杠杆自由现金流").font = FB; sb(ws, r, 1, r, 6, fill=KEY, font=FB)
    for i in range(5):
        cl = CL(2+i)
        fval(ws, r, 2+i, f"={cl}{nopat_dcf}+{cl}{da_dcf}-{cl}{capex_dcf}-{cl}{nwc_dcf}")
    ws.cell(row=r, column=7, value="NOPAT+D&A-CapEx-ΔNWC").font = FNOTE; ws.cell(row=r, column=7).border = BD
    r += 1
    df_dcf = r
    ws.cell(row=r, column=1, value="Discount Factor 折现因子").font = FN; ws.cell(row=r, column=1).border = BD
    for i in range(5):
        fval(ws, r, 2+i, f"=1/(1+Assumptions!$I$18)^{i+1}", "0.0000")
    r += 1
    pv_dcf = r
    ws.cell(row=r, column=1, value="PV of UFCF UFCF现值").font = FB; ws.cell(row=r, column=1).border = BD
    for i in range(5):
        fval(ws, r, 2+i, f"={CL(2+i)}{ufcf_dcf}*{CL(2+i)}{df_dcf}")
    r += 2

    # Terminal Value
    tv_r = r
    ws.cell(row=r, column=1, value="TV 终值 (Gordon Growth Model)").font = FB; ws.cell(row=r, column=1).border = BD
    fval(ws, r, 2, f"=F{ufcf_dcf}*(1+Assumptions!$I$13)/(Assumptions!$I$18-Assumptions!$I$13)")
    ws.cell(row=r, column=7, value="UFCF₂₀₃₀×(1+g)/(WACC-g)").font = FNOTE; ws.cell(row=r, column=7).border = BD
    r += 1
    pvtv_r = r
    ws.cell(row=r, column=1, value="PV of TV 终值现值").font = FB; ws.cell(row=r, column=1).border = BD
    fval(ws, r, 2, f"=B{tv_r}*F{df_dcf}")
    r += 1
    ws.cell(row=r, column=1, value="TV as % of EV TV占比").font = FSM; ws.cell(row=r, column=1).border = BD
    fval(ws, r, 2, f"=B{pvtv_r}/(SUM(B{pv_dcf}:F{pv_dcf})+B{pvtv_r})", NP)
    r += 2

    # EV Bridge
    bridge = [
        ("Sum of PV of UFCF UFCF现值合计", f"=SUM(B{pv_dcf}:F{pv_dcf})", "5年UFCF折现合计"),
        ("PV of TV 终值现值", f"=B{pvtv_r}", ""),
        ("EV 企业价值", None, "=UFCF PV + TV PV"),
        ("(+) Cash & Securities 加回现金", "=Assumptions!$I$21", "年末现金+有价证券"),
        ("(-) Total Debt 减去债务", "=Assumptions!$I$22", "短期+长期债务"),
        ("Equity Value 股权价值", None, "=EV+Cash-Debt"),
        ("Diluted Shares 稀释股数 (mm)", "=Assumptions!$I$23", ""),
        ("Implied Share Price 隐含股价", None, "=Equity Value / Shares"),
    ]
    for i, (name, form, nt) in enumerate(bridge):
        rr = r + i
        ws.cell(row=rr, column=1, value=name).font = FB; ws.cell(row=rr, column=1).border = BD
        if form:
            fval(ws, rr, 2, form)
        elif "EV" in name and "企业" in name:
            fval(ws, rr, 2, f"=B{rr-2}+B{rr-1}")
            sb(ws, rr, 1, rr, 2, fill=SUB, font=FB)
        elif "Equity Value" in name:
            fval(ws, rr, 2, f"=B{rr-3}+B{rr-2}-B{rr-1}")
            sb(ws, rr, 1, rr, 2, fill=SUB, font=FB)
        elif "Implied" in name:
            fval(ws, rr, 2, f"=B{rr-2}/B{rr-1}", NU)
            sb(ws, rr, 1, rr, 2, fill=KEY, font=FB)
        if nt:
            ws.cell(row=rr, column=7, value=nt).font = FNOTE; ws.cell(row=rr, column=7).border = BD

    ws.freeze_panes = "B5"
    add_legend(ws, r + len(bridge) + 3)

    # ═══════════════════════════════════════════════════
    #  TAB 10: SENSITIVITY
    # ═══════════════════════════════════════════════════
    ws = wb.create_sheet("Sensitivity")
    ws["A1"] = "Sensitivity 敏感性分析 - Implied Share Price (USD)"; ws["A1"].font = FT
    ws["A2"] = "WACC 加权平均资本成本 × g 永续增长率"; ws["A2"].font = FSM
    cw(ws, {CL(c): 16 for c in range(1, 9)}); ws.column_dimensions["A"].width = 18

    waccs = [0.075, 0.080, 0.085, 0.090, 0.095, 0.100, 0.105]
    gs = [0.020, 0.025, 0.030, 0.035, 0.040]

    ws.cell(row=4, column=1, value="g \\ WACC").font = FB; ws.cell(row=4, column=1).fill = SUB
    ws.cell(row=4, column=1).border = BD; ws.cell(row=4, column=1).alignment = CT
    for j, w in enumerate(waccs):
        c = ws.cell(row=4, column=2+j, value=w)
        c.number_format = NP2; c.fill = HDR; c.font = FW; c.alignment = CT; c.border = BD

    for i, g in enumerate(gs):
        rr = 5 + i
        ws.cell(row=rr, column=1, value=g).number_format = NP2
        ws.cell(row=rr, column=1).fill = SUB; ws.cell(row=rr, column=1).font = FB
        ws.cell(row=rr, column=1).alignment = CT; ws.cell(row=rr, column=1).border = BD
        for j in range(len(waccs)):
            cl = CL(2+j)
            formula = (
                f"=(SUM(DCF!B{pv_dcf}:F{pv_dcf})"
                f"+(DCF!F{ufcf_dcf}*(1+$A{rr})/({cl}$4-$A{rr}))/((1+{cl}$4)^5)"
                f"+Assumptions!$I$21-Assumptions!$I$22)"
                f"/Assumptions!$I$23"
            )
            ws.cell(row=rr, column=2+j, value=formula).number_format = NU
            ws.cell(row=rr, column=2+j).border = BD

    ws.freeze_panes = "B5"
    add_legend(ws, 12)

    # ═══════════════════════════════════════════════════
    #  TAB 11: RATIO ANALYSIS (NEW)
    # ═══════════════════════════════════════════════════
    ws = wb.create_sheet("Ratio_Analysis")
    ws["A1"] = "Financial Ratio Analysis 财务比率分析"; ws["A1"].font = FT
    ws["A2"] = "单位：百分比或倍数 (% or Multiple)"; ws["A2"].font = FSM
    cw(ws, {CL(c): (40 if c == 1 else (38 if c == NCOL else 14)) for c in range(1, NCOL+1)})
    yh(ws, 3)

    r = 5
    P = "Consolidated_PL"
    S = "Segment_Revenue"
    B = "BS"

    sec(ws, r, "Profitability 盈利能力"); r += 1
    ratio_items_profit = [
        ("Gross Margin 毛利率", f"=({P}!{{c}}{rev_pl}-{P}!{{c}}{R['cogs_pl']})/{P}!{{c}}{rev_pl}", "毛利/收入"),
        ("EBIT Margin 经营利润率", f"={P}!{{c}}{ebit_pl}/{P}!{{c}}{rev_pl}", "经营利润/收入"),
        ("Net Margin 净利率", f"={P}!{{c}}{ni_pl}/{P}!{{c}}{rev_pl}", "净利润/收入"),
        ("ROE 净资产收益率", f"={P}!{{c}}{ni_pl}/{B}!{{c}}{equity_r}", "净利润/股东权益"),
        ("ROA 总资产收益率", f"={P}!{{c}}{ni_pl}/{B}!{{c}}{total_assets_r}", "净利润/总资产"),
    ]
    for name, tmpl, nt in ratio_items_profit:
        ws.cell(row=r, column=1, value=name).font = FB; ws.cell(row=r, column=1).border = BD
        for c in range(2, 10):
            fval(ws, r, c, tmpl.format(c=CL(c)), NP)
        note(ws, r, nt)
        r += 1
    r += 1

    sec(ws, r, "Operating Efficiency 营运效率"); r += 1
    # AR Days
    ws.cell(row=r, column=1, value="AR Days 应收账款周转天数").font = FB; ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        fval(ws, r, c, f"={B}!{CL(c)}{bs_rows['AR']}/{P}!{CL(c)}{rev_pl}*365", "0.0")
    note(ws, r, "应收/收入×365")
    r += 1
    # AP Days
    ws.cell(row=r, column=1, value="AP Days 应付账款周转天数").font = FB; ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        fval(ws, r, c, f"={B}!{CL(c)}{bs_rows['AP']}/{P}!{CL(c)}{R['cogs_pl']}*365", "0.0")
    note(ws, r, "应付/COGS×365")
    r += 1
    # CapEx Intensity
    ws.cell(row=r, column=1, value="CapEx Intensity 资本支出强度").font = FB; ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        fval(ws, r, c, f"=-Cash_Flow!{CL(c)}{capex_cf}/{P}!{CL(c)}{rev_pl}", NP)
    note(ws, r, "|CapEx|/Revenue")
    r += 1
    # CapEx / D&A
    ws.cell(row=r, column=1, value="CapEx / D&A 资本支出/折旧").font = FB; ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        fval(ws, r, c, f"=-Cash_Flow!{CL(c)}{capex_cf}/{P}!{CL(c)}{da_pl}", "0.0x")
    note(ws, r, ">1x表示净投资增长")
    r += 2

    sec(ws, r, "Leverage 杠杆与偿债能力"); r += 1
    ws.cell(row=r, column=1, value="Debt / Assets 资产负债率").font = FB; ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        fval(ws, r, c, f"=({B}!{CL(c)}{total_assets_r}-{B}!{CL(c)}{equity_r})/{B}!{CL(c)}{total_assets_r}", NP)
    note(ws, r, "总负债/总资产")
    r += 1
    ws.cell(row=r, column=1, value="Current Ratio 流动比率").font = FB; ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        fval(ws, r, c, f"={B}!{CL(c)}{total_ca_r}/{B}!{CL(c)}{total_cl_r}", "0.00x")
    note(ws, r, "流动资产/流动负债")
    r += 1
    ws.cell(row=r, column=1, value="Net Cash 净现金 (mm)").font = FB; ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        fval(ws, r, c, f"={B}!{CL(c)}{bs_rows['CashTotal']}-{B}!{CL(c)}{bs_rows['LTDebt']}")
    note(ws, r, "现金+证券-长期债务")
    r += 2

    sec(ws, r, "Cash Flow Quality 现金流质量"); r += 1
    ws.cell(row=r, column=1, value="OCF / Net Income 经营现金流/净利润").font = FB; ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        fval(ws, r, c, f"=Cash_Flow!{CL(c)}{cfo_r}/{P}!{CL(c)}{ni_pl}", "0.00x")
    note(ws, r, "现金流质量，>1为佳")
    r += 1
    ws.cell(row=r, column=1, value="FCF / Revenue 自由现金流率").font = FB; ws.cell(row=r, column=1).border = BD
    for c in range(2, 10):
        fval(ws, r, c, f"=Cash_Flow!{CL(c)}{fcf_r}/{P}!{CL(c)}{rev_pl}", NP)
    note(ws, r, "FCF/Revenue")
    r += 2

    sec(ws, r, "Growth 增长指标"); r += 1
    growth_items = [
        ("Revenue Growth 收入增速", f"{P}!{{c}}{rev_pl}", f"{P}!{{p}}{rev_pl}", "YoY"),
        ("Net Income Growth 净利润增速", f"{P}!{{c}}{ni_pl}", f"{P}!{{p}}{ni_pl}", "YoY"),
        ("EPS Growth 每股收益增速", f"{P}!{{c}}{eps_pl}", f"{P}!{{p}}{eps_pl}", "YoY"),
    ]
    for name, cur_t, prev_t, nt in growth_items:
        ws.cell(row=r, column=1, value=name).font = FB; ws.cell(row=r, column=1).border = BD
        for c in range(3, 10):
            fval(ws, r, c, f"={cur_t.format(c=CL(c))}/{prev_t.format(p=CL(c-1))}-1", NP)
        note(ws, r, nt)
        r += 1
    r += 1

    sec(ws, r, "Key Investment Highlights 关键投资亮点"); r += 1
    highlights = [
        "✓ AI驱动搜索广告持续增长，Search revenue占比>55%，核心现金牛",
        "✓ Google Cloud 48% YoY增长(Q4'25)，利润率从5%→24%快速改善",
        "✓ CapEx高峰期(2026 ~$180B)，AI基础设施投资确立长期竞争优势",
        "✓ 大规模回购+首次派息($0.83/share)，股东回报显著增强",
        "✓ Net cash position >$80B，资产负债表极其稳健",
        "⚠ 风险: AI CapEx回报期不确定、DOJ反垄断诉讼、Cloud竞争加剧",
        "⚠ 风险: 2025 OI&E含$24B非上市证券估值收益(不可持续)",
    ]
    for h in highlights:
        ws.cell(row=r, column=1, value=h).font = FNOTE
        r += 1

    ws.freeze_panes = "B4"
    add_legend(ws, r + 2)

    # ═══════════════════════════════════════════════════
    #  FILL KEY SUMMARY (now all refs available)
    # ═══════════════════════════════════════════════════
    ws = wb["Key_Summary"]
    for i, h in enumerate(["核心财务指标 Key Financial Highlights", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "说明 Notes"]):
        c = ws.cell(row=4, column=1+i, value=h); c.font = FW; c.fill = HDR; c.alignment = CT; c.border = BD

    ks_items = [
        (5, "总收入 Total Revenue", f"Consolidated_PL!{{c}}{rev_pl}", NI, "链接自Consolidated_PL"),
        (6, "  YoY增速 %", None, NP, ""),
        (7, "毛利润 Gross Profit", f"Consolidated_PL!{{c}}{gp_pl}", NI, ""),
        (8, "  毛利率 GPM %", None, NP, ""),
        (9, "营业利润 EBIT", f"Consolidated_PL!{{c}}{ebit_pl}", NI, ""),
        (10, "  营业利润率 OPM %", None, NP, ""),
        (11, "净利润 Net Income", f"Consolidated_PL!{{c}}{ni_pl}", NI, ""),
        (12, "  净利率 NPM %", None, NP, ""),
        (13, "  NI YoY %", None, NP, ""),
        (14, "稀释EPS", f"Consolidated_PL!{{c}}{eps_pl}", NU, ""),
    ]
    for rr, label, tmpl, fmt, nt in ks_items:
        ws.cell(row=rr, column=1, value=label).font = FB; ws.cell(row=rr, column=1).border = BD
        if tmpl:
            for c in range(2, 8):
                yrs = [2023,2024,2025,"2026E","2027E","2028E"]
                # Map summary cols to PL cols: summary col 2-7 → PL col 2-7
                lval(ws, rr, c, f"={tmpl.format(c=CL(c))}", fmt)
        elif "YoY" in label or "yoy" in label.lower():
            for c in range(3, 8):
                fval(ws, rr, c, f"={CL(c)}{rr-1}/{CL(c-1)}{rr-1}-1", NP); ws.cell(row=rr, column=c).font = FSM
        elif "GPM" in label:
            for c in range(2, 8):
                fval(ws, rr, c, f"={CL(c)}7/{CL(c)}5", NP); ws.cell(row=rr, column=c).font = FSM
        elif "OPM" in label:
            for c in range(2, 8):
                fval(ws, rr, c, f"={CL(c)}9/{CL(c)}5", NP); ws.cell(row=rr, column=c).font = FSM
        elif "NPM" in label:
            for c in range(2, 8):
                fval(ws, rr, c, f"={CL(c)}11/{CL(c)}5", NP); ws.cell(row=rr, column=c).font = FSM
        if nt:
            ws.cell(row=rr, column=8, value=nt).font = FNOTE; ws.cell(row=rr, column=8).border = BD

    # Segment revenue
    seg_hdr = 16
    for i, h in enumerate(["分部收入结构 Revenue by Segment", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "说明 Notes"]):
        c = ws.cell(row=seg_hdr, column=1+i, value=h); c.font = FW; c.fill = HDR; c.alignment = CT; c.border = BD

    seg_sum_names = ["Google Services Total", "Google Cloud", "Other Bets"]
    seg_sum_refs = [gs_total_r, R["cloud_rev"], R["ob_rev"]]
    for idx, (nm, ref) in enumerate(zip(seg_sum_names, seg_sum_refs)):
        rr = seg_hdr + 1 + idx * 2
        ws.cell(row=rr, column=1, value=nm).font = FB; ws.cell(row=rr, column=1).border = BD
        for c in range(2, 8):
            lval(ws, rr, c, f"=Segment_Revenue!{CL(c)}{ref}", NI)
        rr2 = rr + 1
        ws.cell(row=rr2, column=1, value="  占比 %").font = FSM; ws.cell(row=rr2, column=1).border = BD
        for c in range(2, 8):
            fval(ws, rr2, c, f"={CL(c)}{rr}/{CL(c)}5", NP); ws.cell(row=rr2, column=c).font = FSM

    # BS highlights
    bs_hdr = 24
    for i, h in enumerate(["资产负债核心指标 Balance Sheet Highlights", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "说明"]):
        c = ws.cell(row=bs_hdr, column=1+i, value=h); c.font = FW; c.fill = HDR; c.alignment = CT; c.border = BD

    bs_sum = [
        (25, "总资产 Total Assets", f"BS!{{c}}{total_assets_r}", NI, ""),
        (26, "股东权益 Total Equity", f"BS!{{c}}{equity_r}", NI, ""),
        (27, "现金及证券 Cash & Securities", f"BS!{{c}}{bs_rows['CashTotal']}", NI, ""),
        (28, "ROE 净资产收益率", None, NP, "净利润/权益"),
        (29, "ROA 总资产收益率", None, NP, "净利润/总资产"),
    ]
    for rr, label, tmpl, fmt, nt in bs_sum:
        ws.cell(row=rr, column=1, value=label).font = FB; ws.cell(row=rr, column=1).border = BD
        if tmpl:
            for c in range(2, 8):
                lval(ws, rr, c, f"={tmpl.format(c=CL(c))}", fmt)
        elif "ROE" in label:
            for c in range(2, 8):
                fval(ws, rr, c, f"={CL(c)}11/{CL(c)}26", NP)
        elif "ROA" in label:
            for c in range(2, 8):
                fval(ws, rr, c, f"={CL(c)}11/{CL(c)}25", NP)
        if nt:
            ws.cell(row=rr, column=8, value=nt).font = FNOTE; ws.cell(row=rr, column=8).border = BD

    # Investment highlights
    rr = 32
    sec_ws = ws
    sec_ws.cell(row=rr, column=1, value="关键投资亮点 Key Investment Highlights").font = FS
    sec_ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=8)
    for c in range(1, 9):
        sec_ws.cell(row=rr, column=c).fill = SUB; sec_ws.cell(row=rr, column=c).border = BD
    rr += 1
    for h in [
        "✓ AI驱动搜索广告核心引擎，Search+YouTube贡献>65%收入",
        "✓ Google Cloud 48% YoY增长(Q4'25)，利润率从5%→24%快速提升",
        "✓ AI CapEx高峰投资($180B/2026)确立长期竞争壁垒",
        "✓ 稳健股东回报：年回购~$50B + 股息$0.88/share",
        "✓ Net cash >$80B，资产负债表极其稳健",
        "⚠ 风险：AI CapEx回报不确定、DOJ反垄断、Cloud竞争激烈",
    ]:
        sec_ws.cell(row=rr, column=1, value=h).font = FNOTE
        rr += 1

    add_legend(ws, rr + 1)

    # ═══════════════════════════════════════════════════
    #  SAVE
    # ═══════════════════════════════════════════════════
    out = os.path.join(r"C:\Users\lxxxxxx\Desktop\谷歌\google_financial_model_2026", "Alphabet_IB_Model_v5.xlsx")
    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    build()
